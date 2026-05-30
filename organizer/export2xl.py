from openpyxl import Workbook, load_workbook


class Export2XL:
    def __init__(self, csv_filename, xl_filename, xl_sheetname):
        self.xl_filename = xl_filename
        self.xl_sheetname = xl_sheetname
        self.data = None
        self.csv_load(csv_filename)

    def csv_load(self, csv_filename):
        self.data = []
        gen = [line.strip() for line in open(csv_filename, 'r') if line[0] != '#']
        for cur_line in gen:
            if cur_line:
                lst = []
                for token in [cl.strip() for cl in cur_line.split(',')]:
                    try:
                        val = float(token)
                    except ValueError:
                        val = token
                    lst.append(val)
                self.data.append(lst)

    def update_xl(self):
        ext = self.xl_filename[-5:].lower()
        keep_vba = False
        if ext == '.xlsm' or ext == '.xltm':
            keep_vba = True

        wb = load_workbook(filename=self.xl_filename, keep_vba=keep_vba)
        try:
            wks = wb[self.xl_sheetname]
        except KeyError:
            # worksheet does not exist, so create a new one
            wb.create_sheet(self.xl_sheetname)
            wks = wb[self.xl_sheetname]

        # export all data into the Excel worksheet
        for row, lst in enumerate(self.data):
            for col, val in enumerate(lst):
                wks.cell(row=row+1, column=col+1).value = val

        wb.save(self.xl_filename)

    def create_xl(self):
        wb = Workbook()
        wks = wb.active
        wks.title = self.xl_sheetname
        for row in self.data:
            wks.append(row)
        wb.save(self.xl_filename)
